# from IPython.display import display
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
# v_combustiveis_csv = pd.read_csv("precos-semestrais-ca-2022-01.csv", sep=';')
v_combustiveis_excel = pd.read_excel("ca-2021-02.xlsx")

# print(pd.options.display.max_rows)
# print(v_combustiveis_csv)
# print(v_combustiveis_excel)
# display(v_combustiveis_excel)
# print(v_combustiveis_excel.head())
# print(v_combustiveis_excel.head(15))
# print(v_combustiveis_excel.shape)
# print(v_combustiveis_excel.shape[0])
# print(v_combustiveis_excel.shape[1])
# print(v_combustiveis_excel.info())
# print(v_combustiveis_excel.describe())
# print(v_combustiveis_excel['Revenda'])
# print(v_combustiveis_excel['Revenda'][0:10])
# v_ca_df = v_combustiveis_excel[[
#     'Revenda', 'Municipio', 'Produto', 'Valor de Venda']]
# # print(v_ca_df)
# print(v_ca_df.loc[0:10])
v_ca_df = v_combustiveis_excel[[
    'Revenda', 'Municipio', 'Produto', 'Valor de Venda']]
v_gas_df = v_ca_df.loc[v_ca_df['Produto'] == 'GASOLINA']
# print(v_gas_df['Valor de Venda'].max())
# print(v_gas_df[['Municipio', 'Revenda', 'Valor de Venda']].max())
# print(v_gas_df[['Municipio', 'Revenda', 'Valor de Venda']])
# print(v_gas_df.sort_values(by='Valor de Venda')[
#       ['Municipio', 'Revenda', 'Valor de Venda']])
# média de preço
# v_municipio = v_combustiveis_excel[v_combustiveis_excel['Bairro']
#                                    == 'BELA VISTA']
# print(v_municipio)
# v_municipio_loc = v_combustiveis_excel.loc[(v_combustiveis_excel['Bairro'] == 'MOOCA') & (
#     v_combustiveis_excel['Municipio'] == 'SAO PAULO')]
# print(v_municipio_loc)
# v_municipio_loc2 = v_combustiveis_excel.loc[(v_combustiveis_excel['Bairro'] == 'MOOCA') & (
#     (v_combustiveis_excel['Produto'] == 'GASOLINA') | (v_combustiveis_excel['Produto'] == 'GASOLINA ADITIVADA'))]
# print(v_combustiveis_excel.loc[(v_combustiveis_excel['Bairro'] == 'MOOCA') & (v_combustiveis_excel['Municipio'] == 'SAO PAULO') & (
#     (v_combustiveis_excel['Produto'] == 'GASOLINA') | (v_combustiveis_excel['Produto'] == 'GASOLINA ADITIVADA')), ['Valor de Venda']].mean())
# v_produto = v_combustiveis_excel[[
#     'Produto', 'Valor de Venda']].groupby(by='Produto').mean().round(2)
# print(v_produto)
# v_combustiveis_excel['Ativo'] = True
# print(v_combustiveis_excel)
# print(v_combustiveis_excel.head(10))

# v_ca_df.to_excel("ca-teste.xlsx", index=False)

# Criar uma coluna "Obs" que tenha nela escrito "MELHOR CIDADE" quando a coluna Municipio for igual a SAO PAULO
# v_combustiveis_excel['Obs'] = ["MELHOR CIDADE" if municipio ==
#                                'SAO PAULO' else None for municipio in v_combustiveis_excel['Municipio']]
# print(v_combustiveis_excel.loc[v_combustiveis_excel['Municipio'].isin(
#     ['SAO PAULO', 'INDAIATUBA', 'CAMPINAS', 'SALTO']), ['Municipio', 'Obs']])

# np.where()

# num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep=";")
# print(num_habitantes_df)

# num_habitantes_df.rename(columns={"Estado": "Estado - Sigla"}, inplace=True)

# num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep=";")
# num_habitantes_df.rename(columns={"Estado": "Estado - Sigla"}, inplace=True)
# print(num_habitantes_df)


# #Agrupar e contar quantos postos tem na cidade..
# postos_por_municipio_df = merge_df.groupby(by=['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
# postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
# postos_por_municipio_df.rename(columns={"Revenda": "Número de Postos"}, inplace=True)
# print(postos_por_municipio_df)

# #AQUI O FINAL ESTÁ COM PROBLEMA

# #Agrupar e contar quantos postos tem na cidade..
# postos_por_municipio_df = merge_df.groupby(by=['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
# print(postos_por_municipio_df.info())
# postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
# postos_por_municipio_df.rename(columns={"Revenda": "NumPostos"}, inplace=True)

# #postos_por_municipio_df['PostosPorHabitante'] = postos_por_municipio_df['NumPostos'] / postos_por_municipio_df['NumHabitantes2021']
# print(postos_por_municipio_df)

# postos_por_municipio_df = merge_df.groupby(
#     by=['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
# postos_por_municipio_df.reset_index(inplace=True)
# # display(postos_por_municipio_df.info())
# postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
# postos_por_municipio_df.rename(columns={"Revenda": "NumPostos"}, inplace=True)

# postos_por_municipio_df['NumHabitantesPorPosto'] = postos_por_municipio_df['NumHabitantes2021'] / \
#     postos_por_municipio_df['NumPostos']
# # display(postos_por_municipio_df.info())
# print(postos_por_municipio_df)


# Vamos brincar de gráficos!!!

# plt.hist(v_combustiveis_excel['Valor de Venda'])

# plt.title("Preço dos combustíveis - Nov/2021")

# plt.xlabel("Preço (em reais)")
# plt.ylabel("Quantidade de Coletas")

# plt.figure(figsize=(7, 5))

c_mean = v_combustiveis_excel['Valor de Venda'].groupby(
    by=v_combustiveis_excel['Produto']).mean()
# print(c_mean)

# # Plotar o gráfico
# c_mean.plot(
#     kind="barh",
#     xlabel="Tipo de Combustível",
#     ylabel="Preço reais/litro",
#     color="red"
# )

# # Grid
# plt.grid()

# # Exibe
# plt.show()
# # Traça a linha vermelha tracejada com o preço médio
# plt.axvline(v_combustiveis_excel['Valor de Venda'].mean(),
#             color='red', linestyle='dashed', linewidth=5)

# plt.show()


# # Vou definir a área do gráfico
# plt.figure(figsize=(7, 5))

# # Plotar o gráfico
# c_mean.plot(
#     kind="barh",
#     xlabel="Tipo de Combustível",
#     ylabel="Preço reais/litro",
#     title="Média de preços por combustível",
#     color="red",
#     alpha=0.3
# )


# plt.grid()

# sns.despine()

# plt.show()

excel = "por_litro.xlsx"
c_mean.to_excel(excel, "Sumário")


wb = load_workbook(excel)

ws = wb['Sumário']

cinza = PatternFill("solid", fgColor="CCCCCC")
ws['A1'].fill = cinza
ws['B1'].fill = cinza
coords = ['A1', 'B1']
for coord in coords:
    ws[coord].fill = cinza


# Onde o preço do combustível for maior ou igual a 6,5 reais (6.5) pinta a fonte
# de vermelho e deixa negrito...
MAX_ROW = ws.max_row
num_linha = 2
while (num_linha <= MAX_ROW):
    coord = 'B'+str(num_linha)  # coord="B{0}".format(num_linha)
    if ws[coord].value >= 6.5:
        ws[coord].font = Font(bold=True, color="FF0000")
    num_linha = num_linha + 1
# Salvar o Excel
wb.save(excel)
